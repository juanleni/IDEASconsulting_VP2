from __future__ import annotations

import asyncio
import hashlib
import io
import json
import re
import secrets
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path

from fastapi import File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from nicegui import app, ui
from openpyxl import load_workbook
from company_context import empresa_id_from_query_for_admin, with_empresa_id

from ideas_utils import enviar_correo_actualizacion_matriz_legal
from legal_matrix_pdf import generar_pdf_reporte_legal_matrix


DB_PATH = 'ideas.db'
DESIGN_HTML_PATH = Path(__file__).resolve().parent / 'data' / 'legal_matrix_design.html'


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_tables() -> None:
    with _connect() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS legal_sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL,
                nombre TEXT NOT NULL, ubicacion TEXT DEFAULT '', jurisdiccion TEXT DEFAULT '',
                actividad TEXT DEFAULT '', activo INTEGER DEFAULT 1, created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS legal_requirements (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL, site_id INTEGER,
                ambito TEXT DEFAULT 'Medio Ambiente', jurisdiccion TEXT DEFAULT 'Nacional',
                organismo TEXT DEFAULT '', tipo_norma TEXT DEFAULT 'Ley', numero TEXT DEFAULT '',
                titulo TEXT NOT NULL, articulo TEXT DEFAULT '', obligacion TEXT DEFAULT '',
                proceso TEXT DEFAULT '', responsable TEXT DEFAULT '', frecuencia TEXT DEFAULT '',
                estado TEXT DEFAULT 'Pendiente', criticidad TEXT DEFAULT 'Media',
                fecha_publicacion TEXT DEFAULT '', fecha_revision TEXT DEFAULT '', proxima_revision TEXT DEFAULT '',
                evidencia_requerida TEXT DEFAULT '', observaciones TEXT DEFAULT '', vigente INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS legal_evidence (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL, requirement_id INTEGER NOT NULL,
                nombre TEXT NOT NULL, archivo_path TEXT DEFAULT '', comentario TEXT DEFAULT '',
                estado_aprobacion TEXT DEFAULT 'Pendiente', cargado_por TEXT DEFAULT '',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS legal_audits (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL, site_id INTEGER,
                fecha TEXT NOT NULL, tipo TEXT DEFAULT 'Anual', auditor TEXT DEFAULT '',
                alcance TEXT DEFAULT '', resultado TEXT DEFAULT 'programada', hallazgos TEXT DEFAULT '',
                plan_accion TEXT DEFAULT '', fecha_cierre TEXT DEFAULT '', created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS legal_alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL, requirement_id INTEGER,
                tipo TEXT DEFAULT 'Actualización normativa', prioridad TEXT DEFAULT 'Media',
                titulo TEXT NOT NULL, detalle TEXT DEFAULT '', estado TEXT DEFAULT 'Nueva',
                fecha TEXT DEFAULT CURRENT_TIMESTAMP, atendida_por TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS legal_audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL, usuario TEXT DEFAULT '',
                accion TEXT NOT NULL, entidad TEXT NOT NULL, entidad_id INTEGER,
                detalle TEXT DEFAULT '', created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS legal_matrix_settings (
                empresa_id INTEGER PRIMARY KEY,
                delete_all_password_hash TEXT
            );
            CREATE TABLE IF NOT EXISTS legal_matrix_updates (
                id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL,
                resumen TEXT NOT NULL, destinatario TEXT DEFAULT '', enviado_por TEXT DEFAULT '',
                enviado_ok INTEGER DEFAULT 0, created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_legal_req_empresa ON legal_requirements(empresa_id);
            CREATE INDEX IF NOT EXISTS idx_legal_log_empresa ON legal_audit_log(empresa_id);
            CREATE INDEX IF NOT EXISTS idx_legal_updates_empresa ON legal_matrix_updates(empresa_id);
        ''')
        cols = {row[1] for row in conn.execute('PRAGMA table_info(legal_matrix_settings)').fetchall()}
        if 'alertas_email_activo' not in cols:
            conn.execute('ALTER TABLE legal_matrix_settings ADD COLUMN alertas_email_activo INTEGER DEFAULT 1')
        if 'alertas_email_dias_anticipacion' not in cols:
            conn.execute('ALTER TABLE legal_matrix_settings ADD COLUMN alertas_email_dias_anticipacion INTEGER DEFAULT 30')
        if 'alertas_email_ultimo_envio' not in cols:
            conn.execute('ALTER TABLE legal_matrix_settings ADD COLUMN alertas_email_ultimo_envio TEXT')


DEFAULT_DELETE_ALL_PASSWORD = 'IDEAS'


def _hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    rounds = 120_000
    digest = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), rounds).hex()
    return f'pbkdf2_sha256${rounds}${salt}${digest}'


def _verify_password(password: str, stored_hash: str) -> bool:
    try:
        _algo, rounds_txt, salt, digest = stored_hash.split('$', 3)
        rounds = int(rounds_txt)
        calc = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), rounds).hex()
        return secrets.compare_digest(calc, digest)
    except Exception:
        return False


def tiene_legal_matrix_password_personalizada(empresa_id: int) -> bool:
    _ensure_tables()
    with _connect() as conn:
        row = conn.execute(
            'SELECT delete_all_password_hash FROM legal_matrix_settings WHERE empresa_id = ?', (empresa_id,)
        ).fetchone()
    return bool(row and row['delete_all_password_hash'])


def verificar_legal_matrix_delete_password(empresa_id: int, password: str) -> bool:
    _ensure_tables()
    password = str(password or '')
    with _connect() as conn:
        row = conn.execute(
            'SELECT delete_all_password_hash FROM legal_matrix_settings WHERE empresa_id = ?', (empresa_id,)
        ).fetchone()
    stored_hash = row['delete_all_password_hash'] if row else None
    if not stored_hash:
        return secrets.compare_digest(password, DEFAULT_DELETE_ALL_PASSWORD)
    return _verify_password(password, stored_hash)


def guardar_legal_matrix_delete_password(empresa_id: int, password: str) -> None:
    _ensure_tables()
    password = (password or '').strip()
    if not password:
        return
    password_hash = _hash_password(password)
    with _connect() as conn:
        conn.execute(
            '''INSERT INTO legal_matrix_settings (empresa_id, delete_all_password_hash) VALUES (?, ?)
               ON CONFLICT(empresa_id) DO UPDATE SET delete_all_password_hash = excluded.delete_all_password_hash''',
            (empresa_id, password_hash),
        )


def obtener_legal_matrix_alert_settings(empresa_id: int) -> dict:
    _ensure_tables()
    with _connect() as conn:
        row = conn.execute(
            '''SELECT alertas_email_activo, alertas_email_dias_anticipacion, alertas_email_ultimo_envio
               FROM legal_matrix_settings WHERE empresa_id = ?''',
            (empresa_id,),
        ).fetchone()
    if not row:
        return {'activo': True, 'dias_anticipacion': 30, 'ultimo_envio': None}
    return {
        'activo': bool(row['alertas_email_activo']) if row['alertas_email_activo'] is not None else True,
        'dias_anticipacion': int(row['alertas_email_dias_anticipacion'] or 30),
        'ultimo_envio': row['alertas_email_ultimo_envio'],
    }


def guardar_legal_matrix_alert_settings(empresa_id: int, activo: bool, dias_anticipacion: int) -> None:
    _ensure_tables()
    dias_anticipacion = max(1, int(dias_anticipacion or 30))
    with _connect() as conn:
        conn.execute(
            '''INSERT INTO legal_matrix_settings (empresa_id, alertas_email_activo, alertas_email_dias_anticipacion)
               VALUES (?, ?, ?)
               ON CONFLICT(empresa_id) DO UPDATE SET
                   alertas_email_activo = excluded.alertas_email_activo,
                   alertas_email_dias_anticipacion = excluded.alertas_email_dias_anticipacion''',
            (empresa_id, 1 if activo else 0, dias_anticipacion),
        )


def obtener_legal_matrix_alertas_vencimiento_abiertas(empresa_id: int) -> list[dict]:
    rows = _rows(
        '''SELECT titulo, detalle, prioridad FROM legal_alerts
           WHERE empresa_id = ? AND estado != 'Cerrada' AND tipo = 'Vencimiento próximo'
           ORDER BY CASE prioridad WHEN 'critica' THEN 0 WHEN 'alta' THEN 1 WHEN 'media' THEN 2 ELSE 3 END''',
        (empresa_id,),
    )
    return [{'titulo': r['titulo'], 'detalle': r['detalle'], 'prioridad': r['prioridad']} for r in rows]


def marcar_legal_matrix_alertas_enviadas(empresa_id: int) -> None:
    with _connect() as conn:
        conn.execute(
            '''INSERT INTO legal_matrix_settings (empresa_id, alertas_email_ultimo_envio) VALUES (?, ?)
               ON CONFLICT(empresa_id) DO UPDATE SET alertas_email_ultimo_envio = excluded.alertas_email_ultimo_envio''',
            (empresa_id, date.today().isoformat()),
        )


def generar_alertas_vencimientos(empresa_id: int, dias_anticipacion: int = 30) -> list[dict]:
    """Crea alertas de 'Vencimiento próximo' para requisitos vigentes cuya próxima revisión
    cae dentro de la ventana indicada (o ya venció), evitando duplicar alertas abiertas
    para el mismo requisito."""
    _ensure_tables()
    limite = (date.today() + timedelta(days=dias_anticipacion)).isoformat()
    hoy = date.today().isoformat()
    reqs = _rows(
        '''SELECT id, titulo, proxima_revision FROM legal_requirements
           WHERE empresa_id = ? AND vigente = 1 AND proxima_revision IS NOT NULL AND proxima_revision != ''
           AND proxima_revision <= ?''',
        (empresa_id, limite),
    )
    if not reqs:
        return []

    abiertas = _rows(
        "SELECT requirement_id FROM legal_alerts WHERE empresa_id = ? AND estado != 'Cerrada' AND requirement_id IS NOT NULL",
        (empresa_id,),
    )
    ids_con_alerta_abierta = {row['requirement_id'] for row in abiertas}

    creadas = []
    for req in reqs:
        if req['id'] in ids_con_alerta_abierta:
            continue
        vencida = req['proxima_revision'] < hoy
        prioridad = 'critica' if vencida else ('alta' if req['proxima_revision'] <= (date.today() + timedelta(days=7)).isoformat() else 'media')
        titulo = f"{'Vencida' if vencida else 'Vence pronto'}: {req['titulo']}"
        detalle = f"Próxima revisión: {req['proxima_revision']}"
        alert_id = _insert('legal_alerts', {
            'empresa_id': empresa_id,
            'requirement_id': req['id'],
            'tipo': 'Vencimiento próximo',
            'prioridad': prioridad,
            'titulo': titulo,
            'detalle': detalle,
            'estado': 'Nueva',
        })
        creadas.append({'id': alert_id, 'titulo': titulo, 'detalle': detalle, 'prioridad': prioridad})
    return creadas


# Cada (accion, entidad) de la bitácora se agrupa bajo un "tipo de cambio" común, ya que
# BAJA (una fila = un requisito) y BAJA MASIVA/BAJA TOTAL (una fila = N requisitos, con la
# cantidad real embebida en el texto del detalle) deben sumar al mismo total de "eliminadas".
_CAMBIO_GRUPOS: dict[tuple[str, str], str] = {
    ('ALTA', 'requisito legal'): 'normas_altas',
    ('MODIFICACIÓN', 'requisito legal'): 'normas_modificadas',
    ('BAJA', 'requisito legal'): 'normas_eliminadas',
    ('BAJA MASIVA', 'requisito legal'): 'normas_eliminadas',
    ('BAJA TOTAL', 'requisito legal'): 'normas_eliminadas',
    ('ALTA', 'sede'): 'sedes_altas',
    ('MODIFICACIÓN', 'sede'): 'sedes_modificadas',
    ('ALTA', 'auditoría'): 'auditorias_altas',
    ('ALTA', 'evidencia'): 'evidencias_altas',
    ('APROBACIÓN', 'evidencia'): 'evidencias_aprobadas',
    ('RESOLUCIÓN', 'alerta'): 'alertas_resueltas',
}
# Claves cuya fila de bitácora representa un lote: la cantidad real viene como número
# al inicio del texto de "detalle" (ej. "42 norma(s) eliminadas..."), no una fila por ítem.
_CAMBIO_CLAVES_LOTE = {('BAJA MASIVA', 'requisito legal'), ('BAJA TOTAL', 'requisito legal')}
_CAMBIO_LABELS: dict[str, tuple[str, str]] = {
    'normas_altas': ('norma nueva agregada', 'normas nuevas agregadas'),
    'normas_modificadas': ('norma actualizada', 'normas actualizadas'),
    'normas_eliminadas': ('norma eliminada', 'normas eliminadas'),
    'sedes_altas': ('sede nueva', 'sedes nuevas'),
    'sedes_modificadas': ('sede actualizada', 'sedes actualizadas'),
    'auditorias_altas': ('auditoría programada', 'auditorías programadas'),
    'evidencias_altas': ('evidencia cargada', 'evidencias cargadas'),
    'evidencias_aprobadas': ('evidencia aprobada', 'evidencias aprobadas'),
    'alertas_resueltas': ('alerta resuelta', 'alertas resueltas'),
}


def generar_resumen_cambios_pendientes(empresa_id: int) -> dict:
    """Arma un punteo automático y breve de los cambios cargados en la matriz legal
    desde la última publicación exitosa al cliente (o desde siempre, si nunca se publicó)."""
    _ensure_tables()
    with _connect() as conn:
        ultima = conn.execute(
            'SELECT created_at FROM legal_matrix_updates WHERE empresa_id = ? AND enviado_ok = 1 ORDER BY id DESC LIMIT 1',
            (empresa_id,),
        ).fetchone()
    desde = ultima['created_at'] if ultima else None
    if desde:
        logs = _rows(
            'SELECT accion, entidad, detalle FROM legal_audit_log WHERE empresa_id = ? AND created_at > ? ORDER BY id ASC',
            (empresa_id, desde),
        )
    else:
        logs = _rows(
            'SELECT accion, entidad, detalle FROM legal_audit_log WHERE empresa_id = ? ORDER BY id ASC',
            (empresa_id,),
        )

    conteos: dict[str, int] = {}
    importadas = 0
    for row in logs:
        clave = (row['accion'], row['entidad'])
        detalle = str(row.get('detalle') or '')
        if clave == ('IMPORTACIÓN', 'matriz legal'):
            match = re.match(r'(\d+)', detalle)
            importadas += int(match.group(1)) if match else 1
            continue
        grupo = _CAMBIO_GRUPOS.get(clave)
        if not grupo:
            continue
        cantidad = 1
        if clave in _CAMBIO_CLAVES_LOTE:
            match = re.match(r'(\d+)', detalle)
            cantidad = int(match.group(1)) if match else 1
        conteos[grupo] = conteos.get(grupo, 0) + cantidad

    puntos: list[str] = []
    if importadas:
        puntos.append(
            f"{importadas} norma{'s' if importadas != 1 else ''} incorporada{'s' if importadas != 1 else ''} "
            'mediante importación de Excel'
        )
    for grupo, (singular, plural) in _CAMBIO_LABELS.items():
        n = conteos.get(grupo)
        if not n:
            continue
        puntos.append(f'{n} {singular if n == 1 else plural}')

    return {'items': puntos, 'since': desde}


def _user_name() -> str:
    return str(app.storage.user.get('session_user_name') or app.storage.user.get('session_user_key') or 'sistema')


def _log(company_id: int, action: str, entity: str, entity_id: int | None, detail: str = '') -> None:
    with _connect() as conn:
        conn.execute(
            'INSERT INTO legal_audit_log (empresa_id, usuario, accion, entidad, entidad_id, detalle) VALUES (?, ?, ?, ?, ?, ?)',
            (company_id, _user_name(), action, entity, entity_id, detail),
        )


def _rows(query: str, params=()) -> list[dict]:
    with _connect() as conn:
        return [dict(row) for row in conn.execute(query, params).fetchall()]


def _insert(table: str, payload: dict) -> int:
    fields = ', '.join(payload)
    marks = ', '.join('?' for _ in payload)
    with _connect() as conn:
        cur = conn.execute(f'INSERT INTO {table} ({fields}) VALUES ({marks})', tuple(payload.values()))
        return int(cur.lastrowid)


def _insert_many(table: str, payloads: list[dict]) -> int:
    if not payloads:
        return 0
    fields = ', '.join(payloads[0])
    marks = ', '.join('?' for _ in payloads[0])
    with _connect() as conn:
        conn.executemany(
            f'INSERT INTO {table} ({fields}) VALUES ({marks})',
            [tuple(payload.values()) for payload in payloads],
        )
    return len(payloads)


def _update(table: str, row_id: int, payload: dict) -> None:
    assignments = ', '.join(f'{field} = ?' for field in payload)
    with _connect() as conn:
        conn.execute(f'UPDATE {table} SET {assignments} WHERE id = ?', (*payload.values(), row_id))


_AUDIT_STATUS_LABELS = {
    'programada': 'Programada', 'en_curso': 'En curso',
    'con_hallazgos': 'Con hallazgos', 'cerrada': 'Cerrada',
}
_PRIORIDAD_ORDEN = {'critica': 0, 'alta': 1, 'media': 2, 'baja': 3}
_PRIORIDAD_LABELS = {'critica': 'Crítica', 'alta': 'Alta', 'media': 'Media', 'baja': 'Baja'}


def _fmt_fecha_es(iso: str | None) -> str:
    if not iso:
        return '-'
    try:
        d = datetime.strptime(str(iso)[:10], '%Y-%m-%d').date()
    except ValueError:
        return str(iso)
    return d.strftime('%d/%m/%Y')


def _fmt_fecha_larga_es(d: date) -> str:
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
             'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    return f'{d.day} de {meses[d.month - 1]} de {d.year}'


def construir_contexto_reporte_legal_matrix(empresa_id: int) -> dict:
    """Arma todos los datos reales (sin inventar) para el reporte ejecutivo PDF."""
    reqs = _rows('SELECT * FROM legal_requirements WHERE empresa_id = ?', (empresa_id,))
    sites = _rows('SELECT * FROM legal_sites WHERE empresa_id = ?', (empresa_id,))
    evidences = _rows('SELECT * FROM legal_evidence WHERE empresa_id = ?', (empresa_id,))
    audits = _rows('SELECT * FROM legal_audits WHERE empresa_id = ? ORDER BY fecha ASC', (empresa_id,))
    alerts = _rows('SELECT * FROM legal_alerts WHERE empresa_id = ?', (empresa_id,))

    site_names = {row['id']: row.get('nombre') or '' for row in sites}
    requirement_site = {row['id']: row.get('site_id') for row in reqs}

    total = len(reqs) or 1
    estados = {'cumple': 0, 'no_cumple': 0, 'pendiente': 0, 'no_aplica': 0}
    for row in reqs:
        estados[_norm_status(row.get('estado'))] += 1
    cumplimiento_pct = round((estados['cumple'] / total) * 100)

    evidencias_pendientes = sum(1 for e in evidences if str(e.get('estado_aprobacion') or '').strip().lower() == 'pendiente')
    alertas_abiertas = [a for a in alerts if str(a.get('estado') or '').strip().lower() != 'cerrada']

    por_area: dict[str, dict] = {}
    for row in reqs:
        area = row.get('ambito') or 'Sin ámbito'
        entry = por_area.setdefault(area, {'total': 0, 'cumple': 0})
        entry['total'] += 1
        if _norm_status(row.get('estado')) == 'cumple':
            entry['cumple'] += 1
    areas = [
        {'area': area, 'requisitos': datos['total'], 'pct': round((datos['cumple'] / datos['total']) * 100)}
        for area, datos in por_area.items()
    ]

    sedes = []
    for site in sites:
        reqs_sede = [r for r in reqs if r.get('site_id') == site['id']]
        cumple_sede = sum(1 for r in reqs_sede if _norm_status(r.get('estado')) == 'cumple')
        pct_sede = round((cumple_sede / len(reqs_sede)) * 100) if reqs_sede else 0
        sedes.append({
            'nombre': site.get('nombre') or '',
            'provincia': site.get('jurisdiccion') or '-',
            'requisitos': len(reqs_sede),
            'pct': pct_sede,
        })

    alertas_ordenadas = sorted(
        alertas_abiertas,
        key=lambda a: (_PRIORIDAD_ORDEN.get(str(a.get('prioridad') or '').lower(), 9), str(a.get('fecha') or '')),
    )
    alertas_view = [{
        'titulo': a.get('titulo') or '',
        'prioridad_code': str(a.get('prioridad') or '').lower() if str(a.get('prioridad') or '').lower() in _PRIORIDAD_LABELS else 'media',
        'prioridad': _PRIORIDAD_LABELS.get(str(a.get('prioridad') or '').lower(), 'Media'),
        'sede': site_names.get(requirement_site.get(a.get('requirement_id')), '-') or '-',
        'fecha': _fmt_fecha_es(a.get('fecha')),
        'responsable': a.get('atendida_por') or 'Sin asignar',
    } for a in alertas_ordenadas]

    auditorias_view = [{
        'fecha': _fmt_fecha_es(a.get('fecha')),
        'sede': site_names.get(a.get('site_id'), '-') or '-',
        'alcance': a.get('alcance') or '-',
        'auditor': a.get('auditor') or '-',
        'estado': _AUDIT_STATUS_LABELS.get(_norm_audit_status(a.get('resultado')), 'Programada'),
    } for a in audits if _norm_audit_status(a.get('resultado')) != 'cerrada']

    hallazgos: list[str] = []
    for row in reqs:
        if _norm_status(row.get('estado')) == 'no_cumple':
            sede = site_names.get(row.get('site_id'), '') or ''
            norma = f"{row.get('tipo_norma') or 'Norma'} {row.get('numero') or ''}".strip()
            hallazgos.append(f"<strong>{row.get('titulo') or ''}</strong> ({norma}) sin cumplimiento en {sede or 'la empresa'}.")
    for e in evidences:
        estado_ev = str(e.get('estado_aprobacion') or '').strip().lower()
        if estado_ev not in {'pendiente', 'rechazado'}:
            continue
        sede = site_names.get(requirement_site.get(e.get('requirement_id')), '') or ''
        etiqueta = 'en revisión pendiente' if estado_ev == 'pendiente' else 'en estado rechazado'
        hallazgos.append(f'Evidencia "{e.get("nombre") or ""}" {etiqueta} ({sede or "sin sede"}).')

    if cumplimiento_pct < 50:
        conclusion = (
            f"El nivel de cumplimiento ({cumplimiento_pct}%) es crítico. Se recomienda una intervención inmediata: "
            f"priorizar el cierre de los {estados['no_cumple']} requisitos no conformes, escalar las "
            f"{len(alertas_abiertas)} alertas abiertas a los responsables de sede y reforzar la gestión documental de evidencias."
        )
    elif cumplimiento_pct < 80:
        conclusion = (
            f"El nivel de cumplimiento ({cumplimiento_pct}%) requiere atención. Se recomienda avanzar en el cierre de los "
            f"{estados['no_cumple']} requisitos no conformes y dar seguimiento a las {len(alertas_abiertas)} alertas abiertas."
        )
    else:
        conclusion = (
            f"El nivel de cumplimiento ({cumplimiento_pct}%) es sólido. Se recomienda mantener el seguimiento periódico "
            f"de los {len(alertas_abiertas)} alertas abiertas y sostener la gestión documental de evidencias."
        )

    return {
        'cumplimiento_pct': cumplimiento_pct,
        'no_cumplidos': estados['no_cumple'],
        'alertas_abiertas': len(alertas_abiertas),
        'evidencias_pendientes': evidencias_pendientes,
        'estados': estados,
        'total_requisitos': len(reqs),
        'areas': areas,
        'sedes': sedes,
        'alertas': alertas_view,
        'auditorias': auditorias_view,
        'hallazgos': hallazgos,
        'conclusion': conclusion,
    }


def go_to_legal_matrix_module(company_id: int | None = None, set_selection_fn=None) -> None:
    if company_id:
        app.storage.user['management_company_id'] = int(company_id)
        if set_selection_fn:
            set_selection_fn(int(company_id), None)
    ui.navigate.to(with_empresa_id('/sistema-gestion/matriz-legal', company_id))


# ---------------------------------------------------------------------------
# Diseño pixel-perfect (claude.ai/design) <-> datos reales de legal_* tables
# ---------------------------------------------------------------------------

def _norm_status(value: str | None) -> str:
    mapping = {
        'cumple': 'cumple', 'no cumple': 'no_cumple', 'no_cumple': 'no_cumple',
        'pendiente': 'pendiente', 'en proceso': 'pendiente',
        'no aplica': 'no_aplica', 'no_aplica': 'no_aplica',
    }
    return mapping.get((value or '').strip().lower(), 'pendiente')


def _norm_criticality(value: str | None) -> str:
    mapping = {'critica': 'critica', 'crítica': 'critica', 'alta': 'alta', 'media': 'media', 'baja': 'baja'}
    return mapping.get((value or '').strip().lower(), 'media')


def _norm_audit_status(value: str | None) -> str:
    mapping = {
        'planificada': 'programada', 'programada': 'programada',
        'en curso': 'en_curso', 'en_curso': 'en_curso',
        'con hallazgos': 'con_hallazgos', 'con_hallazgos': 'con_hallazgos',
        'cerrada': 'cerrada', 'conforme': 'cerrada',
    }
    return mapping.get((value or '').strip().lower(), 'programada')


def _norm_approval(value: str | None) -> str:
    mapping = {'aprobado': 'aprobado', 'rechazado': 'rechazado', 'pendiente': 'pendiente'}
    return mapping.get((value or '').strip().lower(), 'pendiente')


def _requirement_to_design(row: dict, evidence_counts: dict[int, int]) -> dict:
    return {
        'id': str(row['id']),
        'scope': row.get('ambito') or 'Medio Ambiente',
        'jurisdiction': row.get('jurisdiccion') or 'Nacional',
        'agency': row.get('organismo') or '',
        'normType': row.get('tipo_norma') or 'Ley',
        'normNumber': row.get('numero') or '',
        'title': row.get('titulo') or '',
        'obligation': row.get('obligacion') or '',
        'siteId': str(row['site_id']) if row.get('site_id') else None,
        'responsible': row.get('responsable') or '',
        'status': _norm_status(row.get('estado')),
        'criticality': _norm_criticality(row.get('criticidad')),
        'nextReview': row.get('proxima_revision') or '',
        'evidenceCount': evidence_counts.get(row['id'], 0),
    }


def _site_to_design(row: dict, audits_by_site: dict[int, list[str]]) -> dict:
    dates = sorted(d for d in audits_by_site.get(row['id'], []) if d)
    return {
        'id': str(row['id']),
        'name': row.get('nombre') or '',
        'province': row.get('jurisdiccion') or '',
        'municipality': '',
        'address': row.get('ubicacion') or '',
        'activity': row.get('actividad') or '',
        'processes': row.get('actividad') or '',
        'active': bool(row.get('activo', 1)),
        'nextAudit': dates[0] if dates else '-',
    }


def _audit_to_design(row: dict) -> dict:
    return {
        'id': str(row['id']),
        'date': row.get('fecha') or '',
        'siteId': str(row['site_id']) if row.get('site_id') else None,
        'auditor': row.get('auditor') or '',
        'type': row.get('tipo') or 'Interna',
        'scope': row.get('alcance') or '',
        'status': _norm_audit_status(row.get('resultado')),
        'result': '',
        'findings': row.get('hallazgos') or '',
        'actionPlan': row.get('plan_accion') or '',
        'closeDate': row.get('fecha_cierre') or '',
    }


def _evidence_to_design(row: dict, requirement_site: dict[int, int | None]) -> dict:
    req_id = row.get('requirement_id')
    site_id = requirement_site.get(req_id) if req_id else None
    return {
        'id': str(row['id']),
        'name': row.get('nombre') or '',
        'requirementId': str(req_id) if req_id else None,
        'siteId': str(site_id) if site_id else None,
        'type': 'Documento',
        'uploadDate': (row.get('created_at') or '')[:10],
        'expiryDate': '-',
        'user': row.get('cargado_por') or '',
        'status': _norm_approval(row.get('estado_aprobacion')),
        'version': 1,
    }


def _alert_to_design(row: dict) -> dict:
    estado = (row.get('estado') or '').strip().lower()
    resolved = estado in {'cerrada', 'resuelta'}
    read = resolved or estado in {'leida', 'leída', 'atendida'}
    return {
        'id': str(row['id']),
        'type': row.get('tipo') or 'Actualización normativa',
        'priority': _norm_criticality(row.get('prioridad')),
        'title': row.get('titulo') or '',
        'description': row.get('detalle') or '',
        'date': (row.get('fecha') or '')[:10],
        'read': read,
        'assignedTo': row.get('atendida_por') or 'Sin asignar',
        'resolved': resolved,
        'siteId': None,
    }


def _log_to_design(row: dict) -> dict:
    return {
        'id': str(row['id']),
        'timestamp': row.get('created_at') or '',
        'user': row.get('usuario') or '',
        'action': row.get('accion') or '',
        'entity': row.get('entidad') or '',
        'record': str(row.get('entidad_id') or ''),
        'detail': row.get('detalle') or '',
        'ip': '-',
    }


def _is_matrix_admin() -> bool:
    role = str(app.storage.user.get('role') or '')
    local_role = str(app.storage.user.get('local_user_role') or '').strip().upper()
    return role == 'admin' or local_role in {'IDEAS_ADMIN', 'EMPRESA_ADMIN'}


def _is_ideas_staff() -> bool:
    """Solo personal de IDEAS (no el propio cliente) puede publicar actualizaciones al cliente."""
    role = str(app.storage.user.get('role') or '')
    local_role = str(app.storage.user.get('local_user_role') or '').strip().upper()
    return role == 'admin' or local_role == 'IDEAS_ADMIN'


def _build_design_context(empresa_id: int, company_name: str, contact_email: str = '') -> dict:
    reqs = _rows('SELECT * FROM legal_requirements WHERE empresa_id = ?', (empresa_id,))
    sites = _rows('SELECT * FROM legal_sites WHERE empresa_id = ?', (empresa_id,))
    evidences = _rows('SELECT * FROM legal_evidence WHERE empresa_id = ? ORDER BY created_at DESC', (empresa_id,))
    audits = _rows('SELECT * FROM legal_audits WHERE empresa_id = ? ORDER BY fecha DESC', (empresa_id,))
    alerts = _rows('SELECT * FROM legal_alerts WHERE empresa_id = ? ORDER BY fecha DESC', (empresa_id,))
    logs = _rows('SELECT * FROM legal_audit_log WHERE empresa_id = ? ORDER BY id DESC LIMIT 300', (empresa_id,))

    evidence_counts: dict[int, int] = {}
    for evidence in evidences:
        rid = evidence.get('requirement_id')
        if rid:
            evidence_counts[rid] = evidence_counts.get(rid, 0) + 1

    requirement_site = {row['id']: row.get('site_id') for row in reqs}

    audits_by_site: dict[int, list[str]] = {}
    for audit in audits:
        if audit.get('site_id'):
            audits_by_site.setdefault(audit['site_id'], []).append(audit.get('fecha') or '')

    return {
        'company': {'name': company_name, 'cuit': '', 'contactEmail': contact_email},
        'lastNormUpdate': date.today().isoformat(),
        'sites': [_site_to_design(row, audits_by_site) for row in sites],
        'requirements': [_requirement_to_design(row, evidence_counts) for row in reqs],
        'audits': [_audit_to_design(row) for row in audits],
        'evidences': [_evidence_to_design(row, requirement_site) for row in evidences],
        'alerts': [_alert_to_design(row) for row in alerts],
        'auditLog': [_log_to_design(row) for row in logs],
        'realAdmin': _is_matrix_admin(),
        'isIdeasStaff': _is_ideas_staff(),
    }


def _render_design_page(empresa_id: int, company_name: str, contact_email: str = '') -> str:
    template = DESIGN_HTML_PATH.read_text(encoding='utf-8')
    payload = json.dumps(_build_design_context(empresa_id, company_name, contact_email), ensure_ascii=False)
    payload = payload.replace('</', '<\\/')  # evita cerrar el <script> si algún texto contiene "</"
    html = template.replace('__EMPRESA_ID__', str(empresa_id))
    html = html.replace('__DEMO_DATA_JSON__', payload)
    return html


# ---------------------------------------------------------------------------
# Importación de matriz legal desde Excel (plantilla EU4.05.13.00 / Kataster)
# ---------------------------------------------------------------------------

_NIVEL_A_JURISDICCION = {'B': 'Nacional', 'L': 'Provincial', 'K': 'Municipal'}
_ESTADOS_CUMPLE = {'si', 'sí', 'x', 'j'}


def _xlsx_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value) -> str:
    return ' '.join(_xlsx_text(value).lower().split())


def _find_legal_matrix_header(ws) -> tuple[int, dict[str, int]] | None:
    max_col = min(ws.max_column or 30, 30)
    for row in range(1, 13):
        headers = {col: _norm_header(ws.cell(row=row, column=col).value) for col in range(1, max_col + 1)}
        has_cumple = any('cumpli' in text for text in headers.values())
        has_tipo = any(text == 'tipo' for text in headers.values())
        has_nivel = any(text == 'nivel' for text in headers.values())
        if not (has_cumple and has_tipo and has_nivel):
            continue

        def find(*needles: str, exact: bool = False) -> int | None:
            for col, text in headers.items():
                if not text:
                    continue
                if exact:
                    if text in needles:
                        return col
                elif any(needle in text for needle in needles):
                    return col
            return None

        columns = {
            'estado': find('cumpli'),
            'numero': find('abreviatura'),
            'titulo': find('ley', exact=True),
            'organismo': find('de..', 'de.'),
            'fecha': find('último cambio', 'ultimo cambio'),
            'relevancia': find('relevancia'),
            'tipo': find('tipo', exact=True),
            'nivel': find('nivel', exact=True),
            'obligacion': find('información para el usuario', 'informacion para el usuario'),
            'observaciones': find('obligaciones'),
            'responsable_1': find('supervisión', 'supervision', 'checado'),
            'responsable_2': find('responsabilidades'),
            'evidencia': find('documentos', 'registros'),
        }
        if columns['titulo'] and columns['tipo']:
            return row, columns
    return None


def _parse_legal_matrix_excel(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 - cualquier archivo inválido/corrupto
        raise ValueError('No se pudo leer el archivo. Verificá que sea un Excel (.xlsx) válido.') from exc

    header = None
    worksheet = None
    for ws in workbook.worksheets:
        found = _find_legal_matrix_header(ws)
        if found:
            header, worksheet = found, ws
            break
    if not header or worksheet is None:
        raise ValueError(
            'No se reconoció la estructura de la matriz legal (formato EU4.05.13.00). '
            'Verificá que el archivo tenga las columnas "Cumplió", "Tipo" y "Nivel".'
        )
    header_row, cols = header

    def cell(row: int, key: str):
        col = cols.get(key)
        return worksheet.cell(row=row, column=col).value if col else None

    rows: list[dict] = []
    blank_streak = 0
    row = header_row + 1
    max_row = min(worksheet.max_row or header_row, header_row + 5000)
    while row <= max_row and blank_streak < 15:
        titulo_raw = cell(row, 'titulo')
        tipo_raw = cell(row, 'tipo')
        if titulo_raw in (None, '') and tipo_raw in (None, ''):
            blank_streak += 1
            row += 1
            continue
        blank_streak = 0

        titulo = _xlsx_text(titulo_raw)
        if not titulo:
            row += 1
            continue

        nivel = _xlsx_text(cell(row, 'nivel')).upper()
        relevancia = _xlsx_text(cell(row, 'relevancia')).upper()
        estado = _xlsx_text(cell(row, 'estado')).lower()
        fecha_val = cell(row, 'fecha')
        if isinstance(fecha_val, (date, datetime)):
            fecha_publicacion = fecha_val.date().isoformat() if isinstance(fecha_val, datetime) else fecha_val.isoformat()
        else:
            fecha_publicacion = _xlsx_text(fecha_val)
        responsable = ' — '.join(
            part for part in (_xlsx_text(cell(row, 'responsable_1')), _xlsx_text(cell(row, 'responsable_2'))) if part
        )
        organismo = _xlsx_text(cell(row, 'organismo'))
        ambito = 'Salud Ocupacional' if 'trabajo' in organismo.lower() else 'Medio Ambiente'

        rows.append({
            'ambito': ambito,
            'jurisdiccion': _NIVEL_A_JURISDICCION.get(nivel, 'Nacional'),
            'organismo': organismo,
            'tipo_norma': _xlsx_text(cell(row, 'tipo')) or 'Ley',
            'numero': _xlsx_text(cell(row, 'numero')),
            'titulo': titulo,
            'obligacion': _xlsx_text(cell(row, 'obligacion')),
            'responsable': responsable,
            'estado': 'Cumple' if estado in _ESTADOS_CUMPLE else 'Pendiente',
            'criticidad': 'Alta' if relevancia == 'X' else 'Media',
            'fecha_publicacion': fecha_publicacion,
            'evidencia_requerida': _xlsx_text(cell(row, 'evidencia')),
            'observaciones': _xlsx_text(cell(row, 'observaciones')),
        })
        row += 1

    if not rows:
        raise ValueError('No se encontraron filas con datos para importar en el archivo.')
    return rows


def register_legal_matrix_module(ui, deps: dict) -> None:
    _ensure_tables()
    ensure_platform_access = deps['ensure_platform_access']
    shell = deps['shell']
    company_options = deps['company_options']
    current_selection = deps['current_selection']
    set_selection = deps['set_selection']
    obtener_empresa_detalle = deps['obtener_empresa_detalle']

    def _frame_guard(empresa_id: int) -> bool:
        if not app.storage.user.get('platform_auth'):
            return False
        role = str(app.storage.user.get('role') or '')
        if role == 'empresa':
            logged = app.storage.user.get('logged_empresa_id')
            try:
                return int(logged) == int(empresa_id)
            except (TypeError, ValueError):
                return False
        return True

    @ui.page('/sistema-gestion/matriz-legal')
    def legal_matrix_page() -> None:
        if not ensure_platform_access():
            return
        company_map = company_options()
        query_empresa_id = empresa_id_from_query_for_admin()
        company_id = query_empresa_id or app.storage.user.get('management_company_id') or current_selection()[0]
        company_id = int(company_id) if company_id else (next(iter(company_map)) if company_map else None)
        if query_empresa_id and company_id:
            app.storage.user['management_company_id'] = company_id
            set_selection(company_id, None)
        container = shell('Matriz Legal Digital', back_route='/sistema-gestion', module_key='legal_matrix')
        with container:
            with ui.column().classes('w-full gap-3'):
                if not company_id:
                    ui.label('Seleccioná una empresa para comenzar.').classes('text-slate-500')
                    return

                if str(app.storage.user.get('role') or '') == 'admin':
                    with ui.row().classes('w-full justify-end'):
                        selector = ui.select(company_map, value=company_id, label='Empresa').classes('min-w-[280px]').props('outlined dense')

                        def change_company(_e=None) -> None:
                            if selector.value:
                                app.storage.user['management_company_id'] = int(selector.value)
                                set_selection(int(selector.value), None)
                                ui.navigate.to(with_empresa_id('/sistema-gestion/matriz-legal', selector.value))

                        selector.on_value_change(change_company)

                ui.html(
                    f'<iframe src="/legal-matrix-frame/{company_id}" '
                    'style="width:100%;height:calc(100vh - 190px);border:none;border-radius:20px;'
                    'background:#F2F5F9;display:block;" loading="lazy"></iframe>',
                    sanitize=False,
                ).classes('w-full')

    @app.get('/legal-matrix-frame/{empresa_id}')
    def legal_matrix_frame(empresa_id: int):
        if not _frame_guard(empresa_id):
            return HTMLResponse('<p style="font-family:sans-serif;padding:24px;">Acceso no autorizado.</p>', status_code=403)
        empresa = obtener_empresa_detalle(empresa_id) or {}
        company_name = str(empresa.get('razon_social') or company_options().get(empresa_id) or '')
        contact_email = str(empresa.get('contacto_correo') or '').strip()
        return HTMLResponse(_render_design_page(empresa_id, company_name, contact_email))

    @app.get('/api/legal-matrix/{empresa_id}/export-pdf')
    def export_legal_matrix_pdf(empresa_id: int):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        empresa = obtener_empresa_detalle(empresa_id) or {}
        company_name = str(empresa.get('razon_social') or company_options().get(empresa_id) or '')
        contexto = construir_contexto_reporte_legal_matrix(empresa_id)
        pdf_path = generar_pdf_reporte_legal_matrix(
            empresa_id, company_name, _fmt_fecha_larga_es(date.today()), contexto,
        )
        return FileResponse(
            str(pdf_path), media_type='application/pdf',
            headers={'Content-Disposition': f'inline; filename="{pdf_path.name}"'},
        )

    @app.post('/api/legal-matrix/{empresa_id}/requirements')
    async def create_legal_requirement(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        rid = _insert('legal_requirements', {
            'empresa_id': empresa_id,
            'site_id': int(payload['siteId']) if payload.get('siteId') else None,
            'ambito': payload.get('scope') or 'Medio Ambiente',
            'jurisdiccion': payload.get('jurisdiction') or 'Nacional',
            'organismo': payload.get('agency') or '',
            'tipo_norma': payload.get('normType') or 'Ley',
            'numero': payload.get('normNumber') or '',
            'titulo': payload.get('title') or '',
            'obligacion': payload.get('obligation') or '',
            'responsable': payload.get('responsible') or '',
            'estado': payload.get('status') or 'pendiente',
            'criticidad': payload.get('criticality') or 'media',
            'proxima_revision': payload.get('nextReview') or '',
        })
        _log(empresa_id, 'ALTA', 'requisito legal', rid, str(payload.get('title') or ''))
        return JSONResponse({'id': rid})

    @app.post('/api/legal-matrix/{empresa_id}/requirements/import')
    async def import_legal_requirements(empresa_id: int, file: UploadFile = File(...)):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        content = await file.read()

        def _process() -> int:
            rows = _parse_legal_matrix_excel(content)
            for row in rows:
                row['empresa_id'] = empresa_id
            return _insert_many('legal_requirements', rows)

        try:
            imported = await asyncio.to_thread(_process)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        _log(empresa_id, 'IMPORTACIÓN', 'matriz legal', None, f'{imported} normas desde {file.filename}')
        return JSONResponse({'imported': imported})

    @app.put('/api/legal-matrix/{empresa_id}/requirements/{req_id}')
    async def update_legal_requirement(empresa_id: int, req_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        _update('legal_requirements', req_id, {
            'site_id': int(payload['siteId']) if payload.get('siteId') else None,
            'ambito': payload.get('scope') or 'Medio Ambiente',
            'jurisdiccion': payload.get('jurisdiction') or 'Nacional',
            'organismo': payload.get('agency') or '',
            'tipo_norma': payload.get('normType') or 'Ley',
            'numero': payload.get('normNumber') or '',
            'titulo': payload.get('title') or '',
            'obligacion': payload.get('obligation') or '',
            'responsable': payload.get('responsible') or '',
            'estado': payload.get('status') or 'pendiente',
            'criticidad': payload.get('criticality') or 'media',
            'proxima_revision': payload.get('nextReview') or '',
            'updated_at': datetime.now().isoformat(timespec='seconds'),
        })
        _log(empresa_id, 'MODIFICACIÓN', 'requisito legal', req_id, str(payload.get('title') or ''))
        return JSONResponse({'ok': True})

    @app.delete('/api/legal-matrix/{empresa_id}/requirements/{req_id}')
    def delete_legal_requirement(empresa_id: int, req_id: int):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        with _connect() as conn:
            conn.execute('DELETE FROM legal_requirements WHERE id = ? AND empresa_id = ?', (req_id, empresa_id))
        _log(empresa_id, 'BAJA', 'requisito legal', req_id, '')
        return JSONResponse({'ok': True})

    @app.post('/api/legal-matrix/{empresa_id}/requirements/delete-bulk')
    async def delete_legal_requirements_bulk(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        ids = [int(item) for item in (payload.get('ids') or [])]
        if not ids:
            return JSONResponse({'deleted': 0})
        marks = ', '.join('?' for _ in ids)
        with _connect() as conn:
            cur = conn.execute(
                f'DELETE FROM legal_requirements WHERE empresa_id = ? AND id IN ({marks})',
                (empresa_id, *ids),
            )
            deleted = cur.rowcount
        _log(empresa_id, 'BAJA MASIVA', 'requisito legal', None, f'{deleted} norma(s) seleccionada(s)')
        return JSONResponse({'deleted': deleted})

    @app.post('/api/legal-matrix/{empresa_id}/requirements/delete-all')
    async def delete_legal_requirements_all(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        if not _is_matrix_admin():
            raise HTTPException(status_code=403, detail='Solo un administrador puede eliminar toda la matriz legal.')
        payload = await request.json()
        password = str(payload.get('password') or '')
        if not verificar_legal_matrix_delete_password(empresa_id, password):
            raise HTTPException(status_code=401, detail='Contraseña incorrecta.')
        with _connect() as conn:
            cur = conn.execute('DELETE FROM legal_requirements WHERE empresa_id = ?', (empresa_id,))
            deleted = cur.rowcount
        _log(empresa_id, 'BAJA TOTAL', 'requisito legal', None, f'{deleted} norma(s) eliminadas (borrado total)')
        return JSONResponse({'deleted': deleted})

    @app.post('/api/legal-matrix/{empresa_id}/sites')
    async def create_legal_site(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        rid = _insert('legal_sites', {
            'empresa_id': empresa_id,
            'nombre': payload.get('name') or '',
            'ubicacion': payload.get('address') or payload.get('municipality') or '',
            'jurisdiccion': payload.get('province') or '',
            'actividad': payload.get('activity') or payload.get('processes') or '',
        })
        _log(empresa_id, 'ALTA', 'sede', rid, str(payload.get('name') or ''))
        return JSONResponse({'id': rid})

    @app.put('/api/legal-matrix/{empresa_id}/sites/{site_id}')
    async def update_legal_site(empresa_id: int, site_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        _update('legal_sites', site_id, {
            'nombre': payload.get('name') or '',
            'ubicacion': payload.get('address') or payload.get('municipality') or '',
            'jurisdiccion': payload.get('province') or '',
            'actividad': payload.get('activity') or payload.get('processes') or '',
            'activo': 1 if payload.get('active', True) else 0,
        })
        _log(empresa_id, 'MODIFICACIÓN', 'sede', site_id, str(payload.get('name') or ''))
        return JSONResponse({'ok': True})

    @app.post('/api/legal-matrix/{empresa_id}/audits')
    async def create_legal_audit(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        rid = _insert('legal_audits', {
            'empresa_id': empresa_id,
            'site_id': int(payload['siteId']) if payload.get('siteId') else None,
            'fecha': payload.get('date') or date.today().isoformat(),
            'tipo': payload.get('type') or 'Interna',
            'auditor': payload.get('auditor') or '',
            'alcance': payload.get('scope') or '',
            'resultado': 'programada',
        })
        _log(empresa_id, 'ALTA', 'auditoría', rid, str(payload.get('date') or ''))
        return JSONResponse({'id': rid})

    @app.post('/api/legal-matrix/{empresa_id}/evidence')
    async def create_legal_evidence(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        payload = await request.json()
        if not payload.get('requirementId'):
            raise HTTPException(status_code=400, detail='requirementId requerido')
        rid = _insert('legal_evidence', {
            'empresa_id': empresa_id,
            'requirement_id': int(payload['requirementId']),
            'nombre': payload.get('name') or '',
            'cargado_por': payload.get('user') or _user_name(),
        })
        _log(empresa_id, 'ALTA', 'evidencia', rid, str(payload.get('name') or ''))
        return JSONResponse({'id': rid})

    @app.post('/api/legal-matrix/{empresa_id}/evidence/{evidence_id}/approve')
    def approve_legal_evidence(empresa_id: int, evidence_id: int):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        _update('legal_evidence', evidence_id, {'estado_aprobacion': 'aprobado'})
        _log(empresa_id, 'APROBACIÓN', 'evidencia', evidence_id, '')
        return JSONResponse({'ok': True})

    @app.post('/api/legal-matrix/{empresa_id}/alerts/{alert_id}/resolve')
    def resolve_legal_alert(empresa_id: int, alert_id: int):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        _update('legal_alerts', alert_id, {'estado': 'Cerrada'})
        _log(empresa_id, 'RESOLUCIÓN', 'alerta', alert_id, '')
        return JSONResponse({'ok': True})

    @app.get('/api/legal-matrix/{empresa_id}/changes-summary')
    def legal_matrix_changes_summary(empresa_id: int):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        if not _is_ideas_staff():
            raise HTTPException(status_code=403)
        return JSONResponse(generar_resumen_cambios_pendientes(empresa_id))

    @app.post('/api/legal-matrix/{empresa_id}/notify-update')
    async def notify_legal_matrix_update(empresa_id: int, request: Request):
        if not _frame_guard(empresa_id):
            raise HTTPException(status_code=403)
        if not _is_ideas_staff():
            raise HTTPException(status_code=403, detail='Solo el equipo de IDEAS puede publicar actualizaciones al cliente.')
        payload = await request.json()
        resumen = str(payload.get('summary') or '').strip()
        if not resumen:
            raise HTTPException(status_code=400, detail='Ingresá un resumen de la actualización.')
        empresa = obtener_empresa_detalle(empresa_id) or {}
        correo = str(empresa.get('contacto_correo') or '').strip()
        if not correo:
            raise HTTPException(status_code=400, detail='La empresa no tiene un correo de contacto configurado.')
        company_name = str(empresa.get('razon_social') or company_options().get(empresa_id) or '')
        resultado = await asyncio.to_thread(
            enviar_correo_actualizacion_matriz_legal, correo, company_name, resumen, _user_name(),
        )
        _insert('legal_matrix_updates', {
            'empresa_id': empresa_id,
            'resumen': resumen,
            'destinatario': correo,
            'enviado_por': _user_name(),
            'enviado_ok': 1 if resultado.get('ok') else 0,
        })
        _log(empresa_id, 'NOTIFICACIÓN', 'actualización matriz legal', None, resumen)
        return JSONResponse({'ok': bool(resultado.get('ok')), 'to': correo})
